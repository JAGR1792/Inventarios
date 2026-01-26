from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import load_workbook


@dataclass(frozen=True)
class ImportedProduct:
    key: str
    producto: str
    descripcion: str
    unidades: int
    precio_final: float


class ExcelImporter:
    # Header aliases to support multiple Excel formats.
    # - Old format: Producto, Descripcion, unidades, Precio Final
    # - New format: PRODUCTO, PESO, UNIDADES, PRECIO UNITARIO VENTA
    HEADER_ALIASES: dict[str, list[str]] = {
        "producto": ["Producto", "PRODUCTO"],
        "descripcion": ["Descripcion", "DESCRIPCION", "PESO"],
        "unidades": ["unidades", "UNIDADES"],
        "precio_final": [
            "Precio Final",
            "PRECIO FINAL",
            "PRECIO UNITARIO VENTA",
            "PRECIO UNITARIO",
            "PRECIO VENTA",
        ],
    }

    # Used only for error messaging (preferred display names).
    REQUIRED = {
        "producto": "Producto",
        "descripcion": "Descripcion",
        "unidades": "unidades",
        "precio_final": "Precio Final",
    }

    def __init__(
        self,
        xlsx_path: Path,
        worksheet_name: str,
        *,
        engine: str = "openpyxl",
        cache_dir: Path | None = None,
    ):
        self.xlsx_path = xlsx_path
        self.worksheet_name = worksheet_name
        self.engine = (engine or "openpyxl").strip().casefold()
        self.cache_dir = cache_dir
        if self.cache_dir is not None:
            self.cache_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _norm(x: Any) -> str:
        s = str(x or "").strip()
        s = " ".join(s.split())
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        return s.casefold()

    @staticmethod
    def _parse_money(value: Any) -> float:
        if value in (None, ""):
            return 0.0
        if isinstance(value, (int, float)):
            try:
                return float(value)
            except Exception:
                return 0.0

        s = str(value).strip()
        if not s:
            return 0.0

        # Remove currency text/symbols and keep digits/separators.
        s = s.replace("$", "")
        s = s.replace("COP", "")
        s = s.replace("cop", "")
        s = s.strip()
        s = "".join(ch for ch in s if ch.isdigit() or ch in (".", ",", "-"))
        if not s:
            return 0.0

        # Heuristics for thousands/decimal separators.
        if "." in s and "," in s:
            # Common LatAm: 1.234,56
            s = s.replace(".", "")
            s = s.replace(",", ".")
        elif "," in s:
            # Could be 1234,56 or 1,234
            parts = s.split(",")
            if len(parts) == 2 and len(parts[1]) == 2:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "." in s:
            # Could be 1.234 (thousands) or 1234.56 (decimal)
            parts = s.split(".")
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and parts[0].isdigit()):
                s = s.replace(".", "")

        try:
            return float(s)
        except Exception:
            return 0.0

    def _score_row(self, row_vals: list[Any]) -> tuple[int, dict[str, str]]:
        present = {self._norm(v) for v in row_vals if self._norm(v)}
        matched: dict[str, str] = {}
        score = 0
        for field, aliases in self.HEADER_ALIASES.items():
            for a in aliases:
                if self._norm(a) in present:
                    matched[field] = a
                    score += 1
                    break
        return score, matched

    def _price_cache_path(self) -> Path | None:
        if self.cache_dir is None:
            return None
        return self.cache_dir / "price_cache.json"

    def _load_price_cache(self) -> dict[str, float]:
        p = self._price_cache_path()
        if not p or not p.exists():
            return {}
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            return {}
        if not isinstance(data, dict):
            return {}
        out: dict[str, float] = {}
        for k, v in data.items():
            try:
                out[str(k)] = float(v)
            except Exception:
                continue
        return out

    def _save_price_cache(self, cache: dict[str, float]) -> None:
        p = self._price_cache_path()
        if not p:
            return
        tmp = p.with_suffix(".tmp")
        tmp.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(p)

    def _open(self, *, data_only: bool):
        # read_only=True is dramatically faster and avoids huge memory spikes.
        # values_only iteration is used throughout to prevent creating cell objects.
        return load_workbook(
            filename=self.xlsx_path,
            data_only=bool(data_only),
            read_only=True,
        )

    def _find_header(self, ws, scan_rows: int = 40) -> tuple[int, dict[str, int]]:
        best_row = None
        best_score = -1
        best_values = None

        # Iterate values only; ws.max_row can be misleadingly huge if formatting extends.
        max_r = min(int(getattr(ws, "max_row", 0) or 0) or scan_rows, scan_rows)
        for r in range(1, max_r + 1):
            row_vals = list(next(ws.iter_rows(min_row=r, max_row=r, values_only=True)))
            score, _matched = self._score_row(row_vals)
            if score > best_score:
                best_score = score
                best_row = r
                best_values = row_vals

        if best_score < 3 or best_row is None or best_values is None:
            raise RuntimeError(
                "Could not detect header row in the first rows. "
                "Ensure it contains columns like: Producto/PRODUCTO, unidades/UNIDADES, "
                "Precio Final/PRECIO UNITARIO VENTA (and optionally Descripcion/PESO)."
            )

        header_map = {self._norm(name): idx + 1 for idx, name in enumerate(best_values) if self._norm(name)}
        return best_row, header_map

    def _excel_com_read(self) -> list[ImportedProduct]:
        try:
            import win32com.client  # type: ignore
        except Exception as e:
            raise RuntimeError(
                "EXCEL_ENGINE=excel requires pywin32 (Windows). Install it or use EXCEL_ENGINE=openpyxl."
            ) from e

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = None
        try:
            wb = excel.Workbooks.Open(str(self.xlsx_path))
            ws = None
            try:
                ws = wb.Worksheets(self.worksheet_name)
            except Exception:
                try:
                    ws = wb.Worksheets("INVENTARIO")
                except Exception:
                    ws = wb.Worksheets(1)

            used = ws.UsedRange
            values = used.Value
            if values is None:
                return []

            max_r = min(len(values), 30)
            best_row = None
            best_score = -1
            best_values = None
            for r in range(1, max_r + 1):
                row_vals = list(values[r - 1])
                score, _matched = self._score_row(row_vals)
                if score > best_score:
                    best_score = score
                    best_row = r
                    best_values = row_vals

            if best_score < 3 or best_row is None or best_values is None:
                raise RuntimeError("Could not detect header row in Costos")

            header_map = {self._norm(name): idx + 1 for idx, name in enumerate(best_values)}

            def col_any(displays: list[str]) -> tuple[int, str]:
                for display in displays:
                    k = self._norm(display)
                    if k in header_map:
                        return header_map[k], display
                raise RuntimeError(f"Missing required column (any of): {', '.join(displays)}")

            i_prod, _prod_hdr = col_any(self.HEADER_ALIASES["producto"])
            i_units, _units_hdr = col_any(self.HEADER_ALIASES["unidades"])
            i_price, _price_hdr = col_any(self.HEADER_ALIASES["precio_final"])

            # Optional
            try:
                i_desc, desc_hdr = col_any(self.HEADER_ALIASES["descripcion"])
            except Exception:
                i_desc, desc_hdr = -1, ""

            desc_is_peso = self._norm(desc_hdr) == self._norm("PESO")

            out: list[ImportedProduct] = []
            for r in range(best_row + 1, len(values) + 1):
                row_vals = values[r - 1]
                producto = str(row_vals[i_prod - 1] or "").strip()
                if not producto:
                    continue
                descripcion = ""
                if i_desc > 0:
                    descripcion = str(row_vals[i_desc - 1] or "").strip()

                unidades_raw = row_vals[i_units - 1]
                try:
                    unidades = int(float(unidades_raw)) if unidades_raw not in (None, "") else 0
                except (ValueError, TypeError):
                    unidades = 0

                precio_raw = row_vals[i_price - 1]
                precio_final = self._parse_money(precio_raw)

                key = producto
                if desc_is_peso and descripcion:
                    key = f"{producto} - {descripcion}".strip()

                out.append(
                    ImportedProduct(
                        key=key,
                        producto=producto,
                        descripcion=descripcion,
                        unidades=unidades,
                        precio_final=precio_final,
                    )
                )
            return out
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            finally:
                excel.Quit()

    def read_products(self) -> list[ImportedProduct]:
        if not self.xlsx_path.exists():
            raise RuntimeError(f"Excel file not found: {self.xlsx_path}")

        if self.engine == "excel":
            return self._excel_com_read()

        cache = self._load_price_cache()
        cache_changed = False

        wb = self._open(data_only=True)
        # IMPORTANT: if this workbook contains both legacy sheets (e.g. Costos) and the new
        # INVENTARIO sheet, we always prefer INVENTARIO to avoid importing the wrong format.
        inv = next((n for n in wb.sheetnames if self._norm(n) == self._norm("INVENTARIO")), None)
        if inv is not None:
            ws = wb[inv]
        elif self.worksheet_name in wb.sheetnames:
            ws = wb[self.worksheet_name]
        else:
            # Fallback: auto-detect best match in any worksheet.
            best_name = None
            best_score = -1
            for name in wb.sheetnames:
                try:
                    cand = wb[name]
                    header_row, _header_map = self._find_header(cand)
                    row_vals = list(
                        next(cand.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
                    )
                    score, _matched = self._score_row(row_vals)
                    if score > best_score:
                        best_score = score
                        best_name = name
                except Exception:
                    continue

            if best_name is None:
                raise RuntimeError(
                    f"Worksheet '{self.worksheet_name}' not found and no compatible sheet was detected"
                )
            ws = wb[best_name]
        header_row, header_map = self._find_header(ws)

        def col_any(displays: list[str]) -> tuple[int, str]:
            for display in displays:
                k = self._norm(display)
                if k in header_map:
                    return header_map[k], display
            raise RuntimeError(f"Missing required column (any of): {', '.join(displays)}")

        i_prod, _prod_hdr = col_any(self.HEADER_ALIASES["producto"])
        i_units, _units_hdr = col_any(self.HEADER_ALIASES["unidades"])
        i_price, _price_hdr = col_any(self.HEADER_ALIASES["precio_final"])

        # Optional
        try:
            i_desc, desc_hdr = col_any(self.HEADER_ALIASES["descripcion"])
        except Exception:
            i_desc, desc_hdr = -1, ""

        desc_is_peso = self._norm(desc_hdr) == self._norm("PESO")

        out: list[ImportedProduct] = []
        empty_streak = 0
        max_empty_streak = 200  # stop early if the sheet has long blank tails/formatting

        for row_vals in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Guard against short rows
            def at(idx1: int):
                i0 = idx1 - 1
                return row_vals[i0] if i0 >= 0 and i0 < len(row_vals) else None

            producto = str(at(i_prod) or "").strip()
            if not producto:
                empty_streak += 1
                if empty_streak >= max_empty_streak:
                    break
                continue

            empty_streak = 0

            descripcion = ""
            if i_desc > 0:
                descripcion = str(at(i_desc) or "").strip()

            unidades_raw = at(i_units)
            try:
                unidades = int(float(unidades_raw)) if unidades_raw not in (None, "") else 0
            except (ValueError, TypeError):
                unidades = 0

            precio_raw = at(i_price)
            precio_final = self._parse_money(precio_raw)

            key = producto
            if desc_is_peso and descripcion:
                key = f"{producto} - {descripcion}".strip()

            if precio_final == 0.0 and cache.get(key, 0) > 0:
                precio_final = float(cache[key])
            elif precio_final > 0:
                if cache.get(key) != precio_final:
                    cache[key] = float(precio_final)
                    cache_changed = True

            out.append(
                ImportedProduct(
                    key=key,
                    producto=producto,
                    descripcion=descripcion,
                    unidades=unidades,
                    precio_final=precio_final,
                )
            )

        if cache_changed:
            self._save_price_cache(cache)

        # If mostly zeros, try one-time COM read to seed cache (if Excel is installed).
        if out:
            zero_ratio = sum(1 for p in out if float(p.precio_final) == 0.0) / float(len(out))
            if zero_ratio > 0.5:
                try:
                    com = self._excel_com_read()
                    by_key = {p.key: p for p in com if float(p.precio_final) > 0}
                    patched: list[ImportedProduct] = []
                    for p in out:
                        if float(p.precio_final) == 0.0 and p.key in by_key:
                            patched_price = float(by_key[p.key].precio_final)
                            cache[p.key] = patched_price
                            patched.append(
                                ImportedProduct(
                                    key=p.key,
                                    producto=p.producto,
                                    descripcion=p.descripcion,
                                    unidades=p.unidades,
                                    precio_final=patched_price,
                                )
                            )
                        else:
                            patched.append(p)
                    out = patched
                    self._save_price_cache(cache)
                except Exception:
                    pass

        return out
