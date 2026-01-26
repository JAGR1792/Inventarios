from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class ExportRow:
    producto: str
    descripcion: str
    unidades: int
    precio_final: Decimal


HEADERS = ["PRODUCTO", "PESO", "UNIDADES", "PRECIO UNITARIO VENTA"]


def export_inventory_to_excel(*, xlsx_path: Path, worksheet_name: str, rows: list[ExportRow]) -> tuple[int, str]:
    """Overwrite the inventory worksheet with current products.

    This intentionally writes values only (no styles) and preserves other sheets.
    Returns number of data rows written.
    """

    p = Path(xlsx_path).expanduser().resolve()
    if not p.exists():
        raise RuntimeError("El archivo Excel no existe")
    if p.suffix.lower() != ".xlsx":
        raise RuntimeError("El archivo debe ser .xlsx")

    wb = load_workbook(filename=p)
    desired = (worksheet_name or "INVENTARIO").strip() or "INVENTARIO"

    # 1) Prefer matching by name (case-insensitive).
    ws = None
    desired_key = desired.casefold()
    for existing in wb.sheetnames:
        if str(existing).strip().casefold() == desired_key:
            ws = wb[existing]
            break

    # 2) If not found, and user wants INVENTARIO/INVENTARIOS, fall back to the 3rd worksheet (index 2).
    if ws is None and desired_key in ("inventario", "inventarios") and len(wb.worksheets) >= 3:
        ws = wb.worksheets[2]

    # 3) Otherwise create the target sheet.
    if ws is None:
        ws = wb.create_sheet(title=desired)

    # Overwrite values in-place to preserve formatting/tables where possible.
    # Only touch columns A-D (headers + data).
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)

    # Compute last used row in A-D (avoid clearing huge formatted ranges).
    last_used = 1
    for i, vals in enumerate(
        ws.iter_rows(min_row=2, min_col=1, max_col=len(HEADERS), values_only=True),
        start=2,
    ):
        if any(v not in (None, "") for v in vals):
            last_used = i

    write_row = 2
    for r in rows:
        ws.cell(row=write_row, column=1, value=str(r.producto or "").strip())
        ws.cell(row=write_row, column=2, value=str(r.descripcion or "").strip())
        ws.cell(row=write_row, column=3, value=int(r.unidades or 0))
        ws.cell(
            row=write_row,
            column=4,
            value=float(Decimal(str(r.precio_final or 0)).quantize(Decimal("0.01"))),
        )
        write_row += 1

    # Clear remaining old rows (A-D only).
    clear_from = write_row
    clear_to = max(last_used, clear_from - 1)
    if clear_from <= clear_to:
        for rr in range(clear_from, clear_to + 1):
            for cc in range(1, len(HEADERS) + 1):
                ws.cell(row=rr, column=cc, value=None)

    wb.save(p)
    return int(len(rows)), str(ws.title)
